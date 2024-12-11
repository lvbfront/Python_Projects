import openpyxl
from pathlib import Path

excelfile = openpyxl.Workbook()
print(excelfile.sheetnames)

#change sheet name
#excelsheet = excelfile.active
#excelsheet = 'firstSheet'

#creat sheeet
excelfile.create_sheet()
excelfile.create_sheet(index=1, title='secondeSheet')

#delete
del excelfile['Sheet1']

sheet = excelfile['secondeSheet']
sheet['A1'] = 'hello there'



names = ['abdo', 'ahmed', 'hadi', 'cres']
exsheet = excelfile['Sheet']
for i in range(1, len(names) + 1):
    exsheet.cell(row=i, column=3).value = names[i-1]



print(sheet['A1'].value)



#r is important
excelfile.save(filename = Path.home() /Path(r"C:\Users\abdu4\OneDrive\سطح المكتب\python course\excel") / 'new.xlsx')