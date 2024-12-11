import openpyxl
from pathlib import Path


excelFile = openpyxl.load_workbook(Path.home() / Path('C:\\', 'example.xlsx'))      #the excel in the c
print(excelFile.sheetnames)

sheet1 = excelFile['Sheet1']
print(sheet1.title)


activeSheet = excelFile.active
print(activeSheet)

print("---------------")
print(sheet1['A1'].value)
print(sheet1['B1'].value)
print(sheet1['C1'].value)
print(sheet1['C1'].row)  #the position    
print(sheet1['C1'].column)
print(sheet1['C1'].coordinate)

print("-----------")

print(sheet1.cell(row = 1, column=2).value)


for i in range(1, 7):
    print(sheet1.cell(row=i, column=1).value)

for i in range(1, 7):
    print("salary: ", sheet1.cell(row=i, column=2).value)

print("-" * 40)
total = 0
for i in range(1, 7):
    print(sheet1.cell(row=i, column=1).value, sheet1.cell(row=i, column=2).value )
    total += sheet1.cell(row=i, column=2).value

print("total salary of employee: ", total)


print(sheet1.max_column)
print(sheet1.max_row)